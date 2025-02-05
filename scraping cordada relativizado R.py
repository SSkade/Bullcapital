import os
import re
from openpyxl import load_workbook
from openpyxl.utils import FORMULAE

# Obtener el directorio del escritorio del usuario
desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')

# Verificar si existe una carpeta llamada "practicante" en el escritorio
practicante_dir = os.path.join(desktop, "practicante")
if os.path.exists(practicante_dir):
    script_dir = practicante_dir
else:
    script_dir = desktop

# Directorios de los archivos de texto (relativos al directorio del script)
txt_directory = os.path.join(script_dir, "CORDADA", "txts")
# Ruta del archivo Excel de salida (relativa al directorio del script)
excel_path = os.path.join(script_dir, "excel banco", "banco.xlsx")

# Crear los directorios si no existen
if not os.path.exists(txt_directory):
    os.makedirs(txt_directory)
if not os.path.exists(os.path.dirname(excel_path)):
    os.makedirs(os.path.dirname(excel_path))

# Diccionario para almacenar los datos y mapear las columnas
data_keys = {
    "Monto Operación": "Monto",
    "Fecha de Giro": "Fecha",
    "Anticipo": "Financiamiento",
    "Diferencia de Precio": "Dif Precio",
    "Gastos": "Gasto",
    "Tasa": "Tasa"
}

# Función para extraer los datos de un archivo de texto
def extract_data(file_path):
    with open(file_path, "r", encoding="latin-1") as file:
        content = file.read()
    
    extracted_data = {}
    for key in data_keys.keys():
        if key == "Fecha de Giro":
            pattern = rf"{key}\s*:\s*(\d{{2}}/\d{{2}}/\d{{4}})"
        elif key == "Tasa":
            pattern = rf"{key}\s*:\s*([\d.,]+)\s*%"
        elif key == "Anticipo":
            pattern = rf"{key}\s*:\s*([\d.,]+)\s*%"
        else:
            pattern = rf"{key}\s*:\s*(-?\$[\d.,]+)"
        match = re.search(pattern, content)
        extracted_data[key] = match.group(1) if match else None
    return extracted_data

# Cargar el archivo Excel existente
workbook = load_workbook(excel_path)
sheet = workbook.active

# Obtener los índices de las columnas basados en los encabezados
column_indices = {cell.value.strip(): col_idx for col_idx, cell in enumerate(sheet[1], 1)}

# Imprimir los encabezados encontrados en el archivo Excel
print("Encabezados encontrados en el archivo Excel:", list(column_indices.keys()))

# Verificar que todos los encabezados existen en el archivo Excel
for key in data_keys.values():
    if key not in column_indices:
        raise KeyError(f"El encabezado '{key}' no se encuentra en el archivo Excel.")

# Identificar la fila donde está "CORDADA" en la columna "Fondo"
fondo_col_idx = column_indices["Fondo"]
cordada_row = None

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=fondo_col_idx, max_col=fondo_col_idx):
    for cell in row:
        if cell.value == "CORDADA":
            cordada_row = cell.row
            break
    if cordada_row:
        break

if not cordada_row:
    raise ValueError("No se encontró 'CORDADA' en la columna 'Fondo'.")

# Iterar sobre todos los archivos en el directorio de textos
for filename in os.listdir(txt_directory):
    if filename.endswith(".txt"):
        txt_path = os.path.join(txt_directory, filename)
        extracted_data = extract_data(txt_path)
        
        # Añadir la nueva fila con las fórmulas copiadas
        new_row_idx = sheet.max_row + 1
        for col_idx in range(1, sheet.max_column + 1):
            source_cell = sheet.cell(row=cordada_row, column=col_idx)
            target_cell = sheet.cell(row=new_row_idx, column=col_idx)
            if source_cell.has_style:
                target_cell._style = source_cell._style
            if source_cell.data_type == 'f':
                target_cell.value = source_cell.value.replace(str(cordada_row), str(new_row_idx))
            else:
                target_cell.value = source_cell.value
        
        # Escribir los datos en las columnas correspondientes
        for key, value in extracted_data.items():
            col_idx = column_indices[data_keys[key]]
            if key == "Anticipo" and value is not None:
                cell = sheet.cell(row=new_row_idx, column=col_idx)
                cell.value = float(value.replace(',', '.')) / 100  # Convertir a porcentaje
                cell.number_format = '0.0%'  # Formato de porcentaje
            else:
                sheet.cell(row=new_row_idx, column=col_idx).value = value

# Guardar los cambios en el archivo Excel
workbook.save(excel_path)

print("Extracción y guardado en Excel completados.")